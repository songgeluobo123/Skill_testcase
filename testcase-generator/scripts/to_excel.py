#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
to_excel.py — 将结构化测试用例 JSON 转换为 Excel（Step 6 主产物，替代 test-case-export）

10 列：用例编号 | 测试模块 | 用例名称 | 优先级 | 测试类型 | 前置条件 | 测试步骤 | 预期结果 | 适用阶段 | 设计方法
（列定义见 references/output_format.md）

用法：
  python to_excel.py test_cases.json -o "test_output/v1.11 功能测试用例.xlsx"
  python to_excel.py --emit-template "模板.xlsx"        # 仅生成空白模板（含表头与示例行）

输入 JSON 字段（每条用例）：
  module, name(可选), priority(P1-P4), precondition, steps, expected,
  test_data(可选), coverage_rule(可选), design_method(可选, 列表或字符串),
  test_type(可选, 测试类型), stage(可选, 适用阶段), id(可选, 作为用例编号)

依赖：openpyxl
"""

import argparse
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("[错误] 缺少依赖 openpyxl，请先安装: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

COLUMNS = ["用例编号", "测试模块", "用例名称", "优先级", "测试类型", "前置条件", "测试步骤", "预期结果", "适用阶段", "设计方法"]
PRIORITY_SET = {"P1", "P2", "P3", "P4"}


def load_cases(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and "cases" in data:
        return data["cases"]
    if isinstance(data, list):
        return data
    raise ValueError("输入需为用例数组，或含 'cases' 键的对象")


def normalize_priority(p):
    p = (p or "").strip().upper()
    if p == "P0":
        return "P1"  # 旧方案 P0 归一为 P1
    return p


def dm_to_str(dm):
    if not dm:
        return ""
    if isinstance(dm, str):
        return dm
    return " / ".join(dm)


def case_row(c, seq):
    cid = c.get("id") or f"TC-{seq:03d}"
    return [
        cid,
        c.get("module", "未分类"),
        c.get("name") or "",
        normalize_priority(c.get("priority")),
        c.get("test_type") or "功能测试",
        c.get("precondition", ""),
        c.get("steps", ""),
        c.get("expected", ""),
        c.get("stage") or "功能测试",
        dm_to_str(c.get("design_method")),
    ]


def build_workbook(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "功能测试用例"
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for i, c in enumerate(cases, 1):
        ws.append(case_row(c, i))
    widths = [10, 16, 24, 8, 12, 34, 42, 42, 10, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    return wb


def main():
    ap = argparse.ArgumentParser(description="测试用例 JSON → Excel（10 列）")
    ap.add_argument("input", nargs="?", help="用例 JSON 文件路径")
    ap.add_argument("-o", "--output", required=False, help="输出 .xlsx 路径")
    ap.add_argument("--emit-template", metavar="PATH", help="仅生成空白模板到指定路径")
    args = ap.parse_args()

    if args.emit_template:
        wb = Workbook()
        ws = wb.active
        ws.title = "功能测试用例"
        ws.append(COLUMNS)
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="305496")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.append(["TC-001", "示例模块", "正确密码登录", "P1", "功能测试",
                   "1. 用户已注册\n2. 网络正常",
                   "1. 输入手机号\n2. 输入密码\n3. 点击登录", "1. 接受输入\n2. 掩码显示\n3. 跳转首页",
                   "功能测试", "场景法 / 等价类划分"])
        for i, w in enumerate([10, 16, 24, 8, 12, 34, 42, 42, 10, 24], 1):
            ws.column_dimensions[chr(64 + i)].width = w
        wb.save(args.emit_template)
        print(f"[✓] 已生成模板: {args.emit_template}")
        return 0

    if not args.input or not args.output:
        print("[错误] 需提供 input 与 -o（或使用 --emit-template）", file=sys.stderr)
        return 2

    try:
        cases = load_cases(args.input)
    except Exception as e:
        print(f"[错误] 读取用例失败: {e}", file=sys.stderr)
        return 2

    bad = [c.get("id", "?") for c in cases if normalize_priority(c.get("priority")) not in PRIORITY_SET]
    if bad:
        print(f"[警告] 以下用例优先级非法（应为 P1-P4）: {bad[:10]}", file=sys.stderr)

    wb = build_workbook(cases)
    wb.save(args.output)
    print(f"[✓] 已生成 Excel: {args.output}（{len(cases)} 条用例）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
